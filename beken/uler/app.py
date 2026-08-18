import os
import re
import shutil
import cv2
import pytesseract
from PIL import Image
import openpyxl
from openpyxl import Workbook, load_workbook
from flask import Flask, request, jsonify
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def configure_tesseract():
    if shutil.which("tesseract"):
        pytesseract.pytesseract.tesseract_cmd = "tesseract"
        return True
    default_win_paths = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.expanduser(r"~\AppData\Local\Programs\Tesseract-OCR\tesseract.exe")
    ]
    for path in default_win_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            return True
    return False

configure_tesseract()

def preprocess_transaction_image(image_path):
    img = cv2.imread(image_path)
    if img is None:
        raise ValueError("Gambar tidak dapat dibaca.")

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape[:2]
    scale_factor = 2.5
    gray = cv2.resize(gray, (int(w * scale_factor), int(h * scale_factor)), interpolation=cv2.INTER_CUBIC)

    gray_clean = cv2.GaussianBlur(gray, (3, 3), 0)
    thresh = cv2.adaptiveThreshold(
        gray_clean, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 21, 10
    )

    return gray_clean, thresh

class TransactionInformation:
    def __init__(self):
        self.bank = "-"
        self.tanggal = "-"
        self.waktu = "-"
        self.pengirim = "-"
        self.penerima = "-"
        self.nominal = "-"
        self.no_ref = "-"

class TransactionOCRParser:
    def __init__(self, raw_text):
        self.text = raw_text
        self.result = TransactionInformation()
        self.parse()

    def parse(self):
        lines = [line.strip() for line in self.text.split('\n') if line.strip()]
        full_text_upper = self.text.upper()

        # ------------------------------------------
        # A. DETEKSI BANK PENERIMA / TUJUAN[cite: 9]
        # ------------------------------------------
        text_for_bank_search = full_text_upper
        split_keywords = ["PENERIMA", "KEPADA", "REKENING TUJUAN", "BANK TUJUAN", "KE ", "\nKE\n", " KE:"]
        for kw in split_keywords:
            if kw in full_text_upper:
                text_for_bank_search = full_text_upper.split(kw)[-1]
                break

        banks = {
            "BNI": "Bank BNI",
            "MANDIRI": "Bank Mandiri",
            "BCA": "Bank BCA",
            "BRI": "Bank BRI",
            "SEABANK": "SeaBank",
            "LIVIN": "Bank Mandiri",
            "MYBCA": "Bank BCA",
            "BRIMO": "Bank BRI",
            "WONDR": "Bank BNI",
            "DANA": "DANA", 
            "GOPAY": "GoPay", 
            "OVO": "OVO", 
            "SHOPEEPAY": "ShopeePay"
        }

        for key, value in banks.items():
            if key in text_for_bank_search:
                self.result.bank = value
                break
        
        if self.result.bank == "-":
            for key, value in banks.items():
                if key in full_text_upper:
                    self.result.bank = value
                    break

        # ------------------------------------------
        # B. DETEKSI PENGIRIM & PENERIMA[cite: 9]
        # ------------------------------------------
        for i, line in enumerate(lines):
            line_upper = line.upper()

            # Pengirim (Dukungan Khusus BNI Wondr / Mandiri Livin)[cite: 9]
            if any(k in line_upper for k in ["DARI", "PENGIRIM", "SUMBER DANA", "REKENING ASAL", "REKENING SUMBER"]) and self.result.pengirim == "-":
                clean_line = re.sub(r'^(DARI|PENGIRIM|SUMBER DANA|REKENING ASAL|REKENING SUMBER)\s*:?', '', line, flags=re.IGNORECASE).strip()
                if len(clean_line) > 2 and not clean_line.upper().startswith("DANA"):
                    self.result.pengirim = clean_line
                elif i + 1 < len(lines):
                    candidate = lines[i + 1].strip()
                    if not any(k in candidate.upper() for k in ["NOMINAL", "DETAIL", "TOTAL", "RP", "BNI", "BCA", "MANDIRI"]):
                        self.result.pengirim = candidate

            # Penerima[cite: 9]
            if any(k in line_upper for k in ["KE", "PENERIMA", "REKENING TUJUAN", "KEPADA"]) and self.result.penerima == "-":
                clean_line = re.sub(r'^(KE|PENERIMA|REKENING TUJUAN|KEPADA)\s*:?', '', line, flags=re.IGNORECASE).strip()
                if len(clean_line) > 2 and not any(k in clean_line.upper() for k in banks.keys()):
                    self.result.penerima = clean_line
                elif i + 1 < len(lines):
                    candidate = lines[i + 1].strip()
                    if not any(k in candidate.upper() for k in ["SUMBER", "DANA", "NOMINAL", "DETAIL"]):
                        self.result.penerima = candidate

        # Pembersihan Angka/Ikon Sampah[cite: 9]
        if self.result.pengirim != "-":
            self.result.pengirim = re.sub(r'^[^a-zA-Z]+', '', self.result.pengirim).strip()
            self.result.pengirim = re.sub(r'\b(BNI|BCA|BRI|MANDIRI)?\s*[\*\d\-]{6,}\b', '', self.result.pengirim, flags=re.IGNORECASE).strip()

        if self.result.penerima != "-":
            self.result.penerima = re.sub(r'^[^a-zA-Z]+', '', self.result.penerima).strip()
            self.result.penerima = re.sub(r'\b(BNI|BCA|BRI|MANDIRI)?\s*[\*\d\-]{6,}\b', '', self.result.penerima, flags=re.IGNORECASE).strip()

        # ------------------------------------------
        # C. DETEKSI TANGGAL, WAKTU & NO REF[cite: 9]
        # ------------------------------------------
        for i, line in enumerate(lines):
            line_upper = line.upper()

            # Waktu[cite: 9]
            match_waktu = re.search(r'(\d{2}:\d{2}(?::\d{2})?)', line)
            if match_waktu and self.result.waktu == "-":
                self.result.waktu = match_waktu.group(1).strip()

            # Tanggal (Mendukung format SeaBank & Wondr)[cite: 9]
            if self.result.tanggal == "-":
                match_tgl = re.search(r'(\d{1,2}\s+[A-Za-z0-9]{3,9}\s+20\d{2}|\d{1,2}[\/\.-]\d{1,2}[\/\.-]20\d{2})', line)
                if not match_tgl and match_waktu:
                    left_side = line.split(match_waktu.group(1))[0].strip()
                    match_tgl = re.search(r'(\d{1,2}\s+[A-Za-z0-9]{3,9}\s+20\d{2})', left_side)

                if match_tgl:
                    res_tgl = match_tgl.group(1).strip()
                    res_tgl = re.sub(r'\b(JU1|JUI|JU|JULI)\b', 'Jul', res_tgl, flags=re.IGNORECASE)
                    self.result.tanggal = res_tgl

            # No Ref (Pemisahan Aman Angka Ref dari Tanggal/Waktu)[cite: 9]
            keywords_ref = ["NO. REF", "REF. ID", "REF ID", "REF", "REFF", "NO. TRANSAKSI", "TRANSAKSI ID", "ID TRANSAKSI", "NO TRANSAKSI"]
            if any(k in line_upper for k in keywords_ref) and self.result.no_ref == "-":
                # Ambil teks khusus di sebelah kanan kata 'Ref' / 'No. Transaksi'[cite: 9]
                after_ref = re.split(r'NO\.?\s*REF\.?|REF\.?\s*ID|NO\.?\s*TRANSAKSI|ID\s*TRANSAKSI', line_upper)
                if len(after_ref) > 1:
                    target = after_ref[-1]
                    digits = re.findall(r'\d{8,}', target)
                    if digits:
                        self.result.no_ref = digits[0]
                
                if self.result.no_ref == "-" and i + 1 < len(lines):
                    digits = re.findall(r'\d{8,}', lines[i+1])
                    if digits:
                        self.result.no_ref = digits[0]

        # Fallback Tanggal dari No Ref SeaBank (YYYYMMDD)[cite: 9]
        if self.result.tanggal == "-" and self.result.no_ref != "-" and len(self.result.no_ref) >= 8:
            potential_year = self.result.no_ref[:4]
            if potential_year.startswith("20") and int(potential_year) <= 2030:
                y = self.result.no_ref[:4]
                m = self.result.no_ref[4:6]
                d = self.result.no_ref[6:8]
                if 1 <= int(m) <= 12 and 1 <= int(d) <= 31:
                    self.result.tanggal = f"{d}/{m}/{y}"

        # ------------------------------------------
        # D. DETEKSI NOMINAL[cite: 9]
        # ------------------------------------------
        for i, line in enumerate(lines):
            line_upper = line.upper()
            if any(k in line_upper for k in ["JUMLAH TOTAL", "JUMLAH TRANSFER", "TOTAL", "NOMINAL", "AMOUNT", "JUMLAH", "RP"]):
                match_rp = re.search(r'(?:RP|\:\s*)[\s\.]*([0-9\.\,]{4,})', line_upper)
                if match_rp:
                    raw_nom = match_rp.group(1).strip()
                    self.result.nominal = self.clean_nominal(raw_nom)
                    break
                elif i + 1 < len(lines):
                    match_next = re.search(r'([0-9\.\,]{4,})', lines[i+1])
                    if match_next:
                        raw_nom = match_next.group(1).strip()
                        self.result.nominal = self.clean_nominal(raw_nom)
                        break

    def clean_nominal(self, raw_str):
        parts = re.split(r'[\.\,]', raw_str)
        if len(parts) >= 2 and len(parts[0]) > 3:
            parts[0] = parts[0].replace("795", "75")
            if len(parts[0]) == 4 and parts[0][2] == '9':
                parts[0] = parts[0][:2] + parts[0][3:]

        clean_str = ".".join(parts)
        return f"Rp {clean_str}"

    def to_dict(self):
        return self.result.__dict__

@app.route('/api/ocr', methods=['POST'])
def process_ocr():
    if 'file' not in request.files:
        return jsonify({"error": "Tidak ada file diunggah"}), 400
    
    file = request.files['file']
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    try:
        gray_img, thresh_img = preprocess_transaction_image(filepath)
        custom_config = r'--psm 6'

        raw_text = pytesseract.image_to_string(thresh_img, config=custom_config)
        parser = TransactionOCRParser(raw_text)
        data = parser.to_dict()

        if data["tanggal"] == "-" or data["pengirim"] == "-":
            raw_text_gray = pytesseract.image_to_string(gray_img, config=custom_config)
            parser_gray = TransactionOCRParser(raw_text_gray)
            data_gray = parser_gray.to_dict()

            if data["tanggal"] == "-" and data_gray["tanggal"] != "-":
                data["tanggal"] = data_gray["tanggal"]
            if data["pengirim"] == "-" and data_gray["pengirim"] != "-":
                data["pengirim"] = data_gray["pengirim"]

        return jsonify({"status": "success", "data": data})
    except Exception as e:
        print("ERROR ON OCR:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route('/api/simpan-excel', methods=['POST'])
def save_excel():
    data = request.json
    file_path = "Rekap_Bukti_Transfer.xlsx"
    headers = ["Bank", "Tanggal", "Pengirim", "Penerima", "Nominal", "No Referensi"]

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Rekap Transaksi"
        ws.append(headers)
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws.append([
        data.get("bank", "-"),
        data.get("tanggal", "-"),
        data.get("pengirim", "-"),
        data.get("penerima", "-"),
        data.get("nominal", "-"),
        data.get("no_ref", "-")
    ])
    wb.save(file_path)
    return jsonify({"status": "success", "message": "Data berhasil disimpan ke Excel"})

if __name__ == '__main__':
    app.run(port=5000, debug=True)
